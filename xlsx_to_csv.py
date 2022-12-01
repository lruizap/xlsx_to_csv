import csv
from openpyxl import load_workbook


def xlsx_to_csv(File):
    wb = load_workbook(filename=File)
    sheet = wb.active

    csv_data = []
    for value in sheet.iter_rows(values_only=True):
        csv_data.append(list(value))

    with open('E4.csv', 'w') as csv_obj:
        writer = csv.writer(csv_obj, delimiter=',')
        for line in csv_data:
            writer.writerow(line)
